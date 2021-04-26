import logging

from openpyxl import Workbook

from src.utils.excel_loader import ExcelLoader
from src.utils.exist_util import ExistUtil
from src.utils.pinyin import str_to_pinyin


class TransformTool3:

    def __init__(self, workbook: Workbook, worksheet):
        self.workbook = workbook
        self.worksheet = worksheet

    def excute(self) -> Workbook:
        max_column = self.worksheet.max_column  # 最大列数
        for col in range(1, max_column+1):
            pinyin = str_to_pinyin(str(self.worksheet.cell(1, col).value))
            self.worksheet.cell(1, col).value = pinyin

        return self.workbook


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

    filename = input('请输入excel文件名.后缀名: \n')
    sheetname = input('输入工作表名：\n')

    if ExistUtil.check_exists(filename, sheetname):
        excel_loader = ExcelLoader(filename, sheetname)
        workbook, worksheet = excel_loader.load_excel()
        transformTool3 = TransformTool3(workbook, worksheet)
        new_workbook = transformTool3.excute()
        new_workbook.save(filename.replace(".xlsx", "_" + sheetname + "转换3.xlsx"))
        logging.info("转换成功！")
