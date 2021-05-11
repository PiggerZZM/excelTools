import logging

from openpyxl import Workbook

from src.unmerge_tools.unmerge_tool import UnmergeTool
from src.utils.exist_util import ExistUtil
from src.utils.excel_loader import ExcelLoader
from src.utils.str_to_int import str_to_int


class TransformTool1:

    def __init__(self, workbook: Workbook, worksheet, data_row_begin: int,
                 data_row_end: int, data_col_begin: int, data_col_end: int):
        self.workbook = workbook
        self.worksheet = worksheet
        self.data_row_begin = data_row_begin
        self.data_row_end = data_row_end
        self.data_col_begin = data_col_begin
        self.data_col_end = data_col_end

    def __read_and_write(self, worksheet) -> Workbook():
        total_rows = worksheet.max_row
        total_cols = worksheet.max_column
        top_attrs = total_rows - self.data_row_end + self.data_row_begin - 1
        left_attrs = total_cols - self.data_col_end + self.data_col_begin - 1
        logging.info(top_attrs)
        logging.info(left_attrs)
        # 创建新表格
        new_workbook = Workbook()
        new_sheet = new_workbook.create_sheet(worksheet.title)

        # 读取数据并写入到新表格
        new_row = 1
        count = 0
        for row in range(self.data_row_begin, self.data_row_end + 1):
            for col in range(self.data_col_begin, self.data_col_end + 1):
                if worksheet.cell(row, col).value is not None:
                    data = str(worksheet.cell(row, col).value).strip()
                    count += 1
                    for top in range(1, self.data_row_begin):
                        temp = worksheet.cell(top, col).value
                        if temp is not None:
                            new_sheet.cell(new_row, top, str(temp).strip())
                    for left in range(1, self.data_col_begin):
                        temp = worksheet.cell(row, left).value
                        if temp is not None:
                            new_sheet.cell(new_row, left + top_attrs, str(temp).strip())

                    # 写入数值
                    new_sheet.cell(new_row, top_attrs + left_attrs + 1, data)
                    logging.info("数据[{}, {}]写入成功".format(row, col))
                    new_row += 1

        logging.info("数据集有效数据个数：{}".format(count))

        return new_workbook

    def excute(self) -> Workbook:
        unmerge_tool = UnmergeTool(self.worksheet)
        unmerge_tool.excute()
        new_workbook = self.__read_and_write(self.worksheet)
        return new_workbook


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

    filename = input('请输入excel文件名.后缀名: \n')
    sheetname = input("请输入{}需要转换的工作表名称：\n".format(filename))
    print("输入数据集的行列范围")
    data_row_begin = int(input('数据集起始行号：\n'))
    data_row_end = int(input('数据集结束行号：\n'))
    data_col_begin = str_to_int(input('数据集起始列号(英文字母或数字)：\n'))
    data_col_end = str_to_int(input('数据集结束列号(英文字母或数字)：\n'))
    if ExistUtil.check_exists(filename, sheetname):
        excel_loader = ExcelLoader(filename, sheetname)
        workbook, worksheet = excel_loader.load_excel()
        transformTool1 = TransformTool1(workbook, worksheet, data_row_begin, data_row_end, data_col_begin,
                                        data_col_end)
        new_workbook = transformTool1.excute()
        new_workbook.save(filename.replace(".xlsx", "_" + sheetname + "转换1.xlsx"))
        logging.info("转换成功！")
