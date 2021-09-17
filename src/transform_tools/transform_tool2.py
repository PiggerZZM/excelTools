import logging

from src.unmerge_tools.unmerge_tool import UnmergeTool
from openpyxl import Workbook

from src.utils.exist_util import ExistUtil
from src.utils.excel_loader import ExcelLoader


class TransformTool2:
    """
    "只转换上表头"
    """
    def __init__(self, workbook: Workbook, worksheet, begin_row: int, end_row: int):
        self.workbook = workbook
        self.worksheet = worksheet
        self.begin_row = begin_row
        self.end_row = end_row

    def excute(self):
        unmerge_tool = UnmergeTool(self.worksheet)
        unmerge_tool.excute()

        max_column = self.worksheet.max_column  # 最大列数
        previous = ""
        for col in range(1, max_column + 1):
            temp = ""
            for row in range(self.begin_row, self.end_row + 1):
                if previous != str(self.worksheet.cell(row, col).value):
                    previous = str(self.worksheet.cell(row, col).value)
                    temp += previous
            self.worksheet.cell(self.end_row, col).value = temp
        self.worksheet.delete_rows(1, self.end_row - self.begin_row)


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

    filename = input('请输入excel文件名.后缀名: \n')
    sheetname = input('输入工作表名：\n')
    begin_row = int(input("输入表头起始行："))
    end_row = int(input("输入表头结束行："))

    if ExistUtil.check_exists(filename, sheetname):
        excel_loader = ExcelLoader(filename, sheetname)
        workbook, worksheet = excel_loader.load_excel()
        transformTool2 = TransformTool2(workbook, worksheet, begin_row, end_row)
        workbook.save(filename.replace(".xlsx", "_" + sheetname + "转换2.xlsx"))
        logging.info("转换成功！")
