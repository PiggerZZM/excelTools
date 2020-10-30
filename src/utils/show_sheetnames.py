from openpyxl import load_workbook


def show_sheetnames(filename: str):
    workbook = load_workbook(filename)
    return workbook.sheetnames
