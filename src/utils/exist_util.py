import logging
import os

from openpyxl import load_workbook


class ExistUtil:

    @staticmethod
    def check_exists(filename: str, sheetname: str):
        if not os.path.exists(filename):
            logging.warning("excel表格" + filename + "不存在")
            return False
        else:
            workbook = load_workbook(filename)
            if sheetname not in workbook.sheetnames:
                logging.warning("工作表" + sheetname + "不存在")
                return False
        return True
