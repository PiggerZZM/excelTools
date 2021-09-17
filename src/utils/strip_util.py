from openpyxl import Workbook


class StripUtil:

    @staticmethod
    def strip_all_sheets(workbook: Workbook, letter=' '):
        sheet_names = workbook.sheetnames
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            StripUtil.strip(worksheet, letter)

    @staticmethod
    def strip(worksheet, letter=' '):
        total_rows = worksheet.max_row
        total_cols = worksheet.max_column
        for row in range(1, total_rows + 1):
            for col in range(1, total_cols + 1):
                temp = worksheet.cell(row, col).value
                if temp is not None:
                    worksheet.cell(row, col, str(temp).strip(letter))
