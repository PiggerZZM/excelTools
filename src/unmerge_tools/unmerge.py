import logging

from openpyxl import load_workbook
from copy import deepcopy


def unmerge(filename: str):
    workbook = load_workbook(filename)  # 加载excel
    sheet_names = workbook.sheetnames
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]  # 读取工作表

        # 获取所有 合并单元格的 位置信息
        # 是个可迭代对象，单个对象类型：openpyxl.worksheet.cell_range.MultiCellRange
        # print后就是excel坐标信息
        merged_list = deepcopy(worksheet.merged_cells)  # 深拷贝，进行拆分单元格的同时会改变worksheet.merged_cells
        logging.info("检测到{}_{}中合并单元格如下：".format(filename, sheet_name))
        logging.info(merged_list)

        # 拆分合并的单元格 并填充内容
        for merged_area in merged_list:

            # 这里的行和列的起始值（索引），和Excel的一样，从1开始，并不是从0开始（注意）
            r1, r2, c1, c2 = merged_area.min_row, merged_area.max_row, merged_area.min_col, merged_area.max_col

            worksheet.unmerge_cells(start_row=r1, end_row=r2, start_column=c1, end_column=c2)

            # 获取左上角单元格的内容
            first_value = str(worksheet.cell(r1, c1).value).strip()

            # 数据填充
            for row in range(r1, r2 + 1):  # 遍历行
                for col in range(c1, c2 + 1):
                    worksheet.cell(row, col).value = first_value
    return workbook

