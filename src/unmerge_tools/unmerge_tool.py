import logging
from copy import deepcopy

from openpyxl import Workbook

from src.utils.exist_util import ExistUtil
from src.utils.excel_loader import ExcelLoader


class UnmergeTool:

    def __init__(self, workbook: Workbook, worksheet):
        self.workbook = workbook
        self.worksheet = worksheet

    def excute(self):
        logging.info("正在拆分合并单元格，请稍等……")
        self.unmerge()
        return self.workbook, self.worksheet

    def unmerge(self):
        # 获取所有 合并单元格的 位置信息
        # 是个可迭代对象，单个对象类型：openpyxl.worksheet.cell_range.MultiCellRange
        merged_list = deepcopy(self.worksheet.merged_cells)  # 深拷贝，进行拆分单元格的同时会改变worksheet.merged_cells
        logging.info("检测到合并单元格如下：")
        logging.info(merged_list)

        # 拆分合并的单元格 并填充内容
        for merged_area in merged_list:
            r1, r2, c1, c2 = merged_area.min_row, merged_area.max_row, merged_area.min_col, merged_area.max_col
            self.worksheet.unmerge_cells(start_row=r1, end_row=r2, start_column=c1, end_column=c2)
            # 获取左上角单元格的内容
            first_value = str(self.worksheet.cell(r1, c1).value).strip()
            # 数据填充
            for row in range(r1, r2 + 1):  # 遍历行
                for col in range(c1, c2 + 1):
                    self.worksheet.cell(row, col).value = first_value


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    filename = input('请输入excel文件名.后缀名: \n')
    sheetname = input("请输入工作表名：\n")
    if ExistUtil.check_exists(filename, sheetname):
        excel_loader = ExcelLoader(filename, sheetname)
        workbook, worksheet = excel_loader.load_excel()
        unmerge_tool = UnmergeTool(workbook, worksheet)
        new_workbook, worksheet = unmerge_tool.excute()
        new_workbook.save(filename.replace('.xlsx', "_" + sheetname + '_unmerged.xlsx'))
        logging.info("解除合并单元格成功！")
